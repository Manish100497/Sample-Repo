import openpyxl
from openpyxl import load_workbook
import re


wb1=openpyxl.load_workbook('E:\Sequel String\pdf\Hero_PIR\Hero_Hier_Master.xlsx')
sheet_obj=wb1.create_sheet("sheet")
sheet_obj.title='MASTER'


Heading=['BOM HIERARCHY','MASTER','DIRECT SUP','INDIRECT SUPP','FREQUENCY','FROM DATE','TO DATE','PURCHASE GROUP','VALUE','PERCENTAGE','INPUT CURRENCY','OUTPUT CURRENCY','UNIT','FROM CITY','TO CITY','OSP-CONVERSION','OSP-FREIGHT']
sheet_obj.append(Heading)


path='E:\Sequel String\pdf\Hero_PIR\VTV_Affy_Premier.XLSX'
wb=openpyxl.load_workbook(path,data_only=True) #data_only=True to give only value
ws=wb['Sheet1']

path2="E:\Sequel String\pdf\Hero_PIR\AFFY_INDIA_April'21.xlsx"
wb2=openpyxl.load_workbook(path2,data_only=True)
ws1=wb2['April_21']


ls=[]
for i in range(45,467):
    vendor=ws['E'+str(i)].value

    ind_vendor_name = 'AFFY INDIA PVT LTD' 
    if vendor == ind_vendor_name:
        # print(direct_sup)
        direct_sup=ws['A'+str(i)].value
        ind_sup=ws['D'+str(i)].value
        from_date='01.04.2021'
        to_date='31.12.9999'
        frequency='QTLY'
        purchase_group='866'
        part_code=ws['G'+str(i)].value

        for j in range(31,175):
            partcode2=ws1['E'+str(j)].value
            if part_code==partcode2:
                subchild=ws1['AD'+str(j)].value
                # Weight
                Weight=ws1['AE'+str(j)].value
                if Weight==0 or Weight==None:
                    Weight=0                
                Gross_Wgh=float(Weight)
                Gross_Weight=Gross_Wgh/1000

                PartC_subC=part_code+'_'+subchild # Adding parent and subchild part code

                lst = [PartC_subC,'Gross_Weight',direct_sup,ind_sup,frequency,from_date,to_date,purchase_group,Gross_Weight,'','KG'] 
                sheet_obj.append(lst)

                # Conversion Cost
                Conv_Cost=ws1['BA'+str(j)].value
                if Conv_Cost==0 or Conv_Cost==None:
                    Conv_Cost=0
                lst = [part_code,'Conversion_Cost',direct_sup,ind_sup,frequency,from_date,to_date,purchase_group,Conv_Cost,'','INR'] 
                sheet_obj.append(lst)

                #BOP Cost
                BOP_Cost=ws1['AX'+str(j)].value
                # print(BOP_Cost)
                if BOP_Cost==0 or BOP_Cost==None:
                    BOP_Cost=0
                lst = [part_code,'BOP_Cost',direct_sup,ind_sup,frequency,from_date,to_date,purchase_group,BOP_Cost,'','INR'] 
                sheet_obj.append(lst)

                # Production Hour(PHR)
                Pord_hour=ws1['AK'+str(j)].value
                if Pord_hour==0 or Pord_hour==None:
                    Pord_hour=0                                                                           
                lst = [part_code,'Production_Hour',direct_sup,ind_sup,frequency,from_date,to_date,purchase_group,Pord_hour] 
                sheet_obj.append(lst)

                # MHR
                MHR=ws1['AY'+str(j)].value
                if MHR==0 or MHR==None:
                    MHR=0
                lst = [part_code,'MHR',direct_sup,ind_sup,frequency,from_date,to_date,purchase_group,MHR] 
                sheet_obj.append(lst)

                lst = [] 
                sheet_obj.append(lst)

wb1.save("Hero_Hier_Master.xlsx")




