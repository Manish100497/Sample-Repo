import openpyxl
from openpyxl import load_workbook
import re

wb1=openpyxl.Workbook()
# sheet_obj=wb1.create_sheet("sheet")
# sheet_obj.title='HIERARCHY'
sheet_obj=wb1.active
sheet_obj = wb1['Sheet']
sheet_obj.title = 'HIERARCHY'


Heading=['BOM_HIERARCHY','DIRECT SUP','PURCHASING GROUP','INDIRECT SUPP','FROM DATE','TO DATE','Partcode','Partcode Type1(OE)','Partcode Level2','Partcode Type2(RM/BOP/VTV/INM)','Partcode Level3','Partcode Type3(RM/BOP/VTV/INM)','Partcode Level4','Partcode Type4(RM/BOP/VTV/INM)','Partcode Level5','Partcode Type5(RM/BOP/VTV/INM)','Partcode Level6','Partcode Type6(RM/BOP/VTV/INM)']

sheet_obj.append(Heading)


path='E:\Sequel String\pdf\Hero_PIR\VTV_Affy_Premier.XLSX'
wb=openpyxl.load_workbook(path,data_only=True) #data_only=True to give only value
ws=wb['Sheet1']

path2="E:\Sequel String\pdf\Hero_PIR\AFFY_INDIA_April'21.xlsx"
wb2=openpyxl.load_workbook(path2,data_only=True)
ws1=wb2['April_21']


ls=[]

for i in range(45,467):
    vendor=ws['E'+str(i)].value

    in_vendor_name = 'AFFY INDIA PVT LTD'
    if vendor == in_vendor_name:
        # print(direct_sup)
        direct_sup=ws['A'+str(i)].value
        ind_sup=ws['D'+str(i)].value
        from_date='01.04.2021'
        to_date='31.12.9999'
        part_code=ws['G'+str(i)].value
        part_type='OE'
        for j in range(31,175):
            partcode2=ws1['E'+str(j)].value
            if part_code==partcode2:
                subchild=ws1['AD'+str(j)].value
                # print(subchild)
    

        lst = ['',direct_sup,'',ind_sup,from_date,to_date,part_code,part_type]
        lst1=['',direct_sup,'',ind_sup,from_date,to_date,part_code,part_type,subchild,'RM']
        # print(lst)
        sheet_obj.append(lst)
        sheet_obj.append(lst1)
wb1.save("Hero_Hier_Master.xlsx")
